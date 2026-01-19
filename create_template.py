"""
Script to create the base Excel template for month-end reports.
Run this once to generate templates/reports/month_end_template.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ensure templates/reports directory exists
template_dir = "templates/reports"
os.makedirs(template_dir, exist_ok=True)

template_path = os.path.join(template_dir, "month_end_template.xlsx")

wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Define styles (same as current reporter.py)
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
summary_font = Font(bold=True, size=10)
summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Create Transactions sheet with header row (placeholder)
ws_transactions = wb.create_sheet("Transactions")
# Add placeholder header row (will be replaced with actual column names)
header_row = ['Date', 'Description', 'Category', 'Amount', 'Account', 'Reference']
for idx, header in enumerate(header_row, 1):
    cell = ws_transactions.cell(row=1, column=idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Set column widths for placeholder
for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F'], 1):
    ws_transactions.column_dimensions[col_letter].width = 20

# Create Summary sheet with structure
ws_summary = wb.create_sheet("Summary")
# Header row
ws_summary.cell(row=1, column=1, value='Metric').font = header_font
ws_summary.cell(row=1, column=1).fill = header_fill
ws_summary.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws_summary.cell(row=1, column=1).border = border
ws_summary.cell(row=1, column=2, value='Value').font = header_font
ws_summary.cell(row=1, column=2).fill = header_fill
ws_summary.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')
ws_summary.cell(row=1, column=2).border = border

# Set column widths
ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 20

# Create By Category sheet placeholder (optional - will be created if needed)
# Don't create it now - it will be added dynamically if category data exists

# Save template
wb.save(template_path)
print(f"Template created successfully at: {template_path}")
