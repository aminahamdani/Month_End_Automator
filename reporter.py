import requests
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter

def generate_report():
    """
    Fetches transaction data from the local FastAPI server and saves it to an Excel file
    with formatted headers.
    """
    url = "http://127.0.0.1:8000/transactions"
    output_file = "January_Closing_Report.xlsx"

    try:
        print(f"Fetching data from {url}...")
        response = requests.get(url)
        response.raise_for_status()
        
        data = response.json()
        
        # Handle potential error response from the server logic
        if isinstance(data, dict) and "error" in data:
            print(f"Server Error: {data['error']}")
            return

        print("Data received. Converting to DataFrame...")
        df = pd.DataFrame(data)

        print(f"Saving to {output_file}...")
        create_excel_report(df, output_file)

    except requests.exceptions.ConnectionError:
        print("Error: Could not connect to the server. Is main.py running?")
    except requests.exceptions.RequestException as e:
        print(f"HTTP Error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def create_excel_report(df, output_file):
    """
    Generates a styled Excel report from a DataFrame with summary statistics.
    """
    # Calculate summary statistics
    summary_stats = calculate_summary_statistics(df)
    
    # Use a context manager to handle the usage of the ExcelWriter
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write main transaction data
        df.to_excel(writer, index=False, sheet_name='Transactions')
        
        # Access the workbook and sheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Transactions']
        
        # Define styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        summary_font = Font(bold=True, size=10)
        summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply bold formatting and styling to the header row (Row 1)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Apply currency format to 'Amount' column
        # Find the 'Amount' column index (1-based)
        amount_col_idx = None
        for idx, col_name in enumerate(df.columns, 1):
            if col_name == 'Amount':
                amount_col_idx = idx
                break
        
        if amount_col_idx:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=amount_col_idx)
                cell.number_format = '$#,##0.00'
                cell.border = border

        # Apply borders to all data cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                      min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = border

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add visual enhancements: Data bars for Amount column
        if amount_col_idx:
            amount_letter = get_column_letter(amount_col_idx)
            # Apply data bars to visualize amounts (positive/negative)
            data_range = f"{amount_letter}2:{amount_letter}{worksheet.max_row}"
            worksheet.conditional_formatting.add(
                data_range,
                DataBarRule(
                    start_type='percentile', start_value=10,
                    end_type='percentile', end_value=90,
                    color="638EC6",
                    showValue=True, minLength=None, maxLength=None
                )
            )
        
        # Add summary statistics sheet
        summary_df = pd.DataFrame([
            ['Total Transactions', summary_stats['total_count']],
            ['Total Amount', f"${summary_stats['total_amount']:,.2f}"],
            ['Average Amount', f"${summary_stats['average_amount']:,.2f}"],
            ['Maximum Amount', f"${summary_stats['max_amount']:,.2f}"],
            ['Minimum Amount', f"${summary_stats['min_amount']:,.2f}"],
            ['Positive Transactions', summary_stats['positive_count']],
            ['Negative Transactions', summary_stats['negative_count']],
        ], columns=['Metric', 'Value'])
        
        summary_df.to_excel(writer, index=False, sheet_name='Summary')
        summary_sheet = writer.sheets['Summary']
        
        # Style summary sheet
        for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row + 1):
            for cell in row:
                cell.border = border
                if cell.row == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = Font(size=10)
                    if cell.column == 1:
                        cell.font = summary_font
                        cell.fill = summary_fill
        
        # Auto-adjust summary column widths
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 20
        
        # Add category breakdown if available
        category_breakdown = get_category_breakdown(df)
        if category_breakdown is not None:
            category_breakdown.to_excel(writer, index=False, sheet_name='By Category')
            cat_sheet = writer.sheets['By Category']
            
            # Style category sheet
            for row in cat_sheet.iter_rows(min_row=1, max_row=cat_sheet.max_row + 1):
                for cell in row:
                    cell.border = border
                    if cell.row == 1:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.font = Font(size=10)
                        # Format amount columns
                        if cell.column >= 2:  # Amount columns
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '$#,##0.00'
            
            # Auto-adjust category sheet column widths
            for col in cat_sheet.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 25)
                cat_sheet.column_dimensions[column_letter].width = adjusted_width

    print(f"Success! Report saved as {output_file} with summary statistics")


def calculate_summary_statistics(df):
    """
    Calculate summary statistics from the transaction DataFrame.
    
    Args:
        df (pd.DataFrame): The transaction DataFrame
        
    Returns:
        dict: Dictionary containing various statistics
    """
    if 'Amount' not in df.columns:
        return {
            'total_count': len(df),
            'total_amount': 0,
            'average_amount': 0,
            'max_amount': 0,
            'min_amount': 0,
            'positive_count': 0,
            'negative_count': 0
        }
    
    amounts = pd.to_numeric(df['Amount'], errors='coerce')
    valid_amounts = amounts.dropna()
    
    stats = {
        'total_count': len(df),
        'total_amount': valid_amounts.sum() if len(valid_amounts) > 0 else 0,
        'average_amount': valid_amounts.mean() if len(valid_amounts) > 0 else 0,
        'max_amount': valid_amounts.max() if len(valid_amounts) > 0 else 0,
        'min_amount': valid_amounts.min() if len(valid_amounts) > 0 else 0,
        'positive_count': len(valid_amounts[valid_amounts > 0]) if len(valid_amounts) > 0 else 0,
        'negative_count': len(valid_amounts[valid_amounts < 0]) if len(valid_amounts) > 0 else 0
    }
    
    # Add category/grouping analysis if Category or Description column exists
    category_col = None
    for col in ['Category', 'category', 'Type', 'type', 'Description', 'description']:
        if col in df.columns:
            category_col = col
            break
    
    if category_col:
        category_stats = df.groupby(category_col)['Amount'].agg(['sum', 'count', 'mean']).to_dict('index')
        stats['by_category'] = {
            k: {
                'total': float(v['sum']),
                'count': int(v['count']),
                'average': float(v['mean'])
            }
            for k, v in category_stats.items()
        }
    
    return stats


def get_category_breakdown(df):
    """
    Get transaction breakdown by category if available.
    
    Args:
        df (pd.DataFrame): The transaction DataFrame
        
    Returns:
        pd.DataFrame: Category breakdown or None if no category column
    """
    category_col = None
    for col in ['Category', 'category', 'Type', 'type', 'Description', 'description']:
        if col in df.columns:
            category_col = col
            break
    
    if not category_col or 'Amount' not in df.columns:
        return None
    
    breakdown = df.groupby(category_col)['Amount'].agg([
        ('Total', 'sum'),
        ('Count', 'count'),
        ('Average', 'mean'),
        ('Min', 'min'),
        ('Max', 'max')
    ]).reset_index()
    
    breakdown.columns = ['Category', 'Total Amount', 'Transaction Count', 'Average Amount', 'Min Amount', 'Max Amount']
    return breakdown

if __name__ == "__main__":
    generate_report()
